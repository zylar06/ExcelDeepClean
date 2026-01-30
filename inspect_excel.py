
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing

file_path = 'test_data.xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Workbook loaded: {file_path}")
    print(f"Sheet names: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nScanning sheet: {sheet_name}")
        
        # Check for images/drawings
        if hasattr(ws, '_images') and ws._images:
            print(f"  Found {len(ws._images)} images/drawings.")
            for idx, img in enumerate(ws._images):
                print(f"    Image {idx+1}: {type(img)}")
        else:
            print("  No images found via _images.")

        # Check for legacy drawing objects if possible (openpyxl support varies)
        
        # Scan first few rows for text that might look like buttons
        print("  Scanning first 20 rows for potential button-like text (e.g., 'Edit', 'Delete', 'Save'):")
        for row in ws.iter_rows(max_row=20, values_only=True):
            # Check if any cell content looks like a button label
            row_str = [str(cell).strip() if cell else "" for cell in row]
            potential_buttons = [c for c in row_str if c in ['Edit', 'Delete', 'Save', 'Modify', 'View', '操作', '编辑', '删除']]
            if potential_buttons:
                print(f"    Row contains potential button text: {potential_buttons}")
            
            # Also print the row if it contains 'Button' or similar
            if any("Button" in c for c in row_str):
                 print(f"    Row contains 'Button': {row_str}")

except Exception as e:
    print(f"Error inspecting Excel file: {e}")
