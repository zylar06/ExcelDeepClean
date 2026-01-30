import argparse
import sys
import os
import logging
from .core.reader import StreamReader
from .core.detector import TableDetector
from .core.extractor import TableExtractor
from .core.writers import TableWriter
from .core.excel_writer import ExcelWriter
from .ai.processor import AIProcessor
import json
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

def setup_logging(verbose=False):
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

def main():
    parser = argparse.ArgumentParser(description="Extract structured tables from Excel files.")
    
    subparsers = parser.add_subparsers(dest="command", help="Command to run")
    
    # Extract Command
    extract_parser = subparsers.add_parser("extract", help="Extract tables from Excel")
    extract_parser.add_argument("input_file", help="Path to input .xlsx file")
    extract_parser.add_argument("--output", "-o", default="output", help="Output directory")
    extract_parser.add_argument("--format", "-f", choices=['json', 'csv'], default='json', help="Output format")
    extract_parser.add_argument("--verbose", "-v", action="store_true", help="Enable verbose logging")
    
    # Process JSON Command
    process_parser = subparsers.add_parser("process-json", help="Refine extracted JSON with AI")
    process_parser.add_argument("input_json", help="Path to extracted tables.json")
    process_parser.add_argument("--output", "-o", required=True, help="Output Excel file path (.xlsx)")
    process_parser.add_argument("--api-key", help="LLM API Key")
    process_parser.add_argument("--base-url", default="https://api.deepseek.com/v1", help="LLM Base URL")
    process_parser.add_argument("--verbose", "-v", action="store_true", help="Enable verbose logging")

    args = parser.parse_args()
    
    if not args.command:
        # Default to extract if no command (backward compatibility)
        # But we need to handle if input_file arg is present in root
        if hasattr(args, 'input_file'):
            args.command = 'extract'
        else:
            parser.print_help()
            sys.exit(1)

    setup_logging(args.verbose)
    logger = logging.getLogger("cli")

    if args.command == "extract":
        run_extract(args, logger)
    elif args.command == "process-json":
        run_process_json(args, logger)

def run_process_json(args, logger):
    if not os.path.exists(args.input_json):
        logger.error(f"Input JSON not found: {args.input_json}")
        sys.exit(1)
        
    try:
        logger.info(f"Loading JSON from {args.input_json}...")
        with open(args.input_json, 'r', encoding='utf-8') as f:
            tables = json.load(f)
            
        processor = AIProcessor(api_key=args.api_key, base_url=args.base_url)
        writer = ExcelWriter(args.output)
        
        final_tables = []
        full_audit_log = []
        
        logger.info(f"Processing {len(tables)} tables with AI...")
        for table in tables:
            processed_subtables, log = processor.process_table(table)
            final_tables.extend(processed_subtables)
            full_audit_log.extend(log)
            logger.info(f"  Table {table.get('table_id')}: Split into {len(processed_subtables)} tables, {len(log)} audit actions.")
            
        logger.info(f"Writing {len(final_tables)} tables to {args.output}...")
        writer.write(final_tables, full_audit_log)
        logger.info("Done.")
        
    except Exception as e:
        logger.error(f"Processing failed: {e}", exc_info=True)
        sys.exit(1)

def run_extract(args, logger):
    if not os.path.exists(args.input_file):
        logger.error(f"Input file not found: {args.input_file}")
        sys.exit(1)
        
    try:
        logger.info(f"Processing {args.input_file}...")
        reader = StreamReader(args.input_file)
        
        # Determine sheets to process
        # StreamReader doesn't list sheets directly yet, but we can access internal wb if loaded, or use openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(args.input_file, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        detector = TableDetector()
        extractor = TableExtractor()
        writer = TableWriter(args.output, args.format)
        
        all_extracted_tables = []
        
        for sheet in sheet_names:
            logger.info(f"Analyzing sheet: {sheet}")
            
            # 1. Detect
            candidates = detector.detect(reader, sheet)
            logger.info(f"  Found {len(candidates)} candidates")
            
            # 2. Extract
            # Note: We should probably create a new reader or ensure reader can be reset/reused efficiently
            # StreamReader.iter_sheet handles re-opening if wb is closed? 
            # My implementation of StreamReader keeps _wb open. 
            # But iter_rows in read_only mode can usually be called multiple times? 
            # Documentation says: "If you want to iterate through the cells of a worksheet multiple times, you must use a fresh workbook object."
            # So my StreamReader needs to close and re-open _wb for each pass if we iterate multiple times?
            # detector.detect iterates once.
            # extractor.extract iterates once.
            # So we iterate the sheet TWICE.
            # We must close and re-open.
            
            reader.close()
            # StreamReader doesn't auto-reopen in my current impl if _wb is set.
            # I should modify StreamReader to handle reset or just instantiate new one.
            # Instantiating new one is safer.
            
            reader_for_extract = StreamReader(args.input_file)
            # Use cached merged cells to avoid re-parsing
            reader_for_extract._merged_cells_cache = reader.merged_cells
            
            tables = list(extractor.extract_all(reader_for_extract, candidates))
            logger.info(f"  Extracted {len(tables)} tables")
            
            all_extracted_tables.extend(tables)
            reader_for_extract.close()
            
        # 3. Write
        logger.info(f"Writing {len(all_extracted_tables)} tables to {args.output}...")
        writer.write(all_extracted_tables)
        logger.info("Done.")
        
    except Exception as e:
        logger.error(f"Extraction failed: {e}", exc_info=True)
        sys.exit(1)
    finally:
        if 'reader' in locals():
            reader.close()
