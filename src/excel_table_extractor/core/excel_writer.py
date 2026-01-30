import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import logging
import os

class ExcelWriter:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.logger = logging.getLogger("excel_writer")

    def write(self, tables: List[Dict[str, Any]], audit_log: List[Dict[str, Any]]):
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        # Write tables
        for i, table in enumerate(tables):
            sheet_name = table.get('sheet', f"Table_{i}")
            # Sanitize sheet name (limit 31 chars, no special chars)
            sheet_name = self._sanitize_sheet_name(sheet_name)[:31]
            
            # If exists, append suffix
            count = 1
            original_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name[:28]}_{count}"
                count += 1
                
            ws = wb.create_sheet(sheet_name)
            self._write_table_to_sheet(ws, table)
            
        # Write Audit Log
        if audit_log:
            ws_audit = wb.create_sheet("Audit_Log")
            headers = ["original_table_id", "row_index", "action", "reason", "content"]
            ws_audit.append(headers)
            for entry in audit_log:
                ws_audit.append([
                    entry.get("original_table_id"),
                    entry.get("row_index"),
                    entry.get("action"),
                    entry.get("reason"),
                    entry.get("content")
                ])
                
        # Save
        # Ensure directory exists
        os.makedirs(os.path.dirname(os.path.abspath(self.output_path)), exist_ok=True)
        wb.save(self.output_path)
        self.logger.info(f"Saved Excel report to {self.output_path}")

    def _write_table_to_sheet(self, ws, table: Dict[str, Any]):
        columns = table.get('columns', [])
        rows = table.get('rows', [])
        
        # Write Header
        ws.append(columns)
        
        # Write Data
        for row in rows:
            # row is dict, need to map to list based on columns
            row_list = []
            for col in columns:
                val = row.get(col)
                # Convert list/dict to string if needed
                if isinstance(val, (list, dict)):
                    val = str(val)
                row_list.append(val)
            ws.append(row_list)
            
    def _sanitize_sheet_name(self, name: str) -> str:
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        return name
