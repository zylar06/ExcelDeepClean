import openpyxl
from typing import Generator, List, Any, Dict, Tuple, Optional
from ..utils.xlsx_utils import get_merged_cells

class StreamReader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self._merged_cells_cache = None
        self._wb = None

    @property
    def merged_cells(self) -> Dict[str, List[Tuple[int, int, int, int]]]:
        if self._merged_cells_cache is None:
            self._merged_cells_cache = get_merged_cells(self.file_path)
        return self._merged_cells_cache

    def iter_sheet(self, sheet_name: str) -> Generator[Tuple[int, List[Any]], None, None]:
        """
        Yields (row_idx, row_values) with merged cells filled.
        row_idx is 1-based.
        """
        # Ensure workbook is loaded
        if not self._wb:
            self._wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        
        if sheet_name not in self._wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name} not found")
            
        ws = self._wb[sheet_name]
        
        # Get merged ranges for this sheet
        sheet_ranges = self.merged_cells.get(sheet_name, [])
        
        # Optimization: Sort ranges by min_row
        # Range format: (min_col, min_row, max_col, max_row)
        sheet_ranges.sort(key=lambda x: x[1])
        
        # Active ranges: ranges that span across the current row
        # We need to store the value of the top-left cell for each active range
        # Key: range_tuple, Value: cell_value
        active_values: Dict[Tuple[int, int, int, int], Any] = {}
        
        # Pointer to the next range to consider from the sorted list
        next_range_idx = 0
        total_ranges = len(sheet_ranges)
        
        # Helper to find if a cell is in an active range
        def get_enclosing_range(r, c, active_ranges):
            for rng in active_ranges:
                min_c, min_r, max_c, max_r = rng
                if min_r <= r <= max_r and min_c <= c <= max_c:
                    return rng
            return None

        for row in ws.iter_rows(values_only=True):
            # openpyxl iter_rows(values_only=True) yields a tuple of values
            # But we don't get the row index easily if there are gaps? 
            # Actually read_only iter_rows yields rows sequentially.
            # We need to track row index manually or use enumerate.
            pass
        
        # Re-implementing loop to get row index correctly
        # In read_only, ws.iter_rows() yields Row objects or values.
        # If we use values_only=False, we get Cell objects which have .row attribute.
        # But values_only=True is faster. 
        # However, empty rows might be skipped or return None?
        # Let's use enumerate(ws.iter_rows(values_only=True), start=1) assuming standard behavior.
        # Note: openpyxl read_only might skip empty rows at the start but usually yields None for empty cells.
        
        current_row_idx = 0
        
        # Active ranges for the current row
        current_active_ranges = []
        
        for row_cells in ws.iter_rows(values_only=True):
            current_row_idx += 1
            
            # Update active ranges
            # 1. Remove ranges that ended before this row
            current_active_ranges = [
                rng for rng in current_active_ranges 
                if rng[3] >= current_row_idx # max_row >= current
            ]
            
            # 2. Add new ranges starting at this row
            while next_range_idx < total_ranges:
                rng = sheet_ranges[next_range_idx]
                min_col, min_row, max_col, max_row = rng
                
                if min_row == current_row_idx:
                    current_active_ranges.append(rng)
                    next_range_idx += 1
                elif min_row < current_row_idx:
                    # Should not happen if sorted, unless ranges overlap weirdly or we missed some logic
                    next_range_idx += 1
                else:
                    # min_row > current_row_idx, stop adding
                    break
            
            if not current_active_ranges:
                yield current_row_idx, list(row_cells)
                continue

            # Process cells in this row
            new_row_values = []
            for col_idx, value in enumerate(row_cells, start=1):
                # Check if this cell belongs to any active range
                enclosing_range = None
                
                # fast check: is col_idx within any active range's col bounds?
                for rng in current_active_ranges:
                    if rng[0] <= col_idx <= rng[2]:
                        enclosing_range = rng
                        break
                
                if enclosing_range:
                    min_col, min_row, max_col, max_row = enclosing_range
                    
                    if min_row == current_row_idx and min_col == col_idx:
                        # This is the top-left cell. Store value.
                        active_values[enclosing_range] = value
                        new_row_values.append(value)
                    else:
                        # This is a merged cell. Use stored value.
                        val = active_values.get(enclosing_range)
                        new_row_values.append(val)
                else:
                    new_row_values.append(value)
            
            yield current_row_idx, new_row_values
            
            # Cleanup active_values for ranges that end on this row
            # Actually we can do this lazily or periodically.
            # To keep memory low, let's remove values for ranges that are no longer active in the *next* row.
            # But wait, we filtered `current_active_ranges` at the START of the loop.
            # So `active_values` might hold data for ranges that just finished.
            # Let's clean up `active_values` based on `current_active_ranges`?
            # No, `current_active_ranges` is for the *current* row.
            # We can clean up values for ranges where `max_row == current_row_idx`.
            
            keys_to_remove = [
                rng for rng in list(active_values.keys())
                if rng[3] <= current_row_idx
            ]
            for k in keys_to_remove:
                del active_values[k]

    def close(self):
        if self._wb:
            self._wb.close()
